#!/usr/bin/env python3
"""
import_panama_real_control.py — Carga el control real de Agustín del rollout
Arcos Dorados Panamá (PA-1000066) al 25-abr-2026.

Fuente única de verdad:
    "Hoja de control Panama/McDonalds_Control_ Panama_25-04-26.xlsx"

Diferencia vs seed_arcos_claro.py:
- seed_arcos_claro carga 89 sites teóricos del SOW + 2 demo WOs
- este script carga las 89 visitas reales que Agustín ejecutó (70 completed + 17
  cancelled + 3 NA) con handshakes (arrival/start/end) + assets + asset_events
  reales (serials Cisco únicos)

Granularidad:
- 1 sitio físico = Loc.Code base (sin sufijo K1/K2/K3)
  Ej: P22 = un site, no 4 sites distintos (P22 + P22K1 + P22K2 + P22K3)
- 1 visita del tech = 1 work_order
  Refs como "after the previous" en PlannedTime indican mismo WO con +1 asset
- 1 equipo Cisco instalado = 1 asset + 1 asset_event tipo "installed"

Idempotente: re-ejecutable sin duplicar (upsert por keys naturales).

Pre-requisitos:
- seed_arcos_claro.py ya corrido (necesita CES org, agreement V1.1, project)
- Usuario agustinc@systemrapid.com existe (seed_foundation lo crea)

Usage:
    docker compose exec api python -m scripts.import_panama_real_control [--dry-run]
"""
from __future__ import annotations

import argparse
import asyncio
import re
import sys
from datetime import datetime, time, timedelta, timezone
from pathlib import Path

# Add project root to path
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from app.database import connect_db, get_db  # noqa: E402

# Excel control source — versioned copy inside scripts/_data/
# (the api container only mounts ./backend, so we keep a copy here)
# Original lives at: <repo_root>/Hoja de control Panama/McDonalds_Control_ Panama_25-04-26.xlsx
XLSX_PATH = Path(__file__).resolve().parent / "_data" / "panama_control_25-04-26.xlsx"

PROJECT_CODE = "ARCOS-CLARO-SDWAN-OFFNET"
SERVICE_AGREEMENT_REF = "04MSP-V1.1"
TECH_EMAIL = "agustinc@systemrapid.com"
COORD_EMAIL = "androsb@systemrapid.com"
SOW_SITES_JSON = Path(__file__).resolve().parent / "_panama_sites.json"
PANAMA_STATUS_ORDER = ["intake", "triage", "pre_flight", "dispatched"]


def load_sow_lookup() -> dict[str, dict]:
    """Índice loc_base → {lat, lng, name, fm_ids} desde la lista SOW (_panama_sites.json).

    La hoja de control usa loc codes (P22, P22K1, 101, 106...). La lista SOW usa
    nombres tipo "PANAMA P07 PLAZA LA MITRA Restaurante" / "PAN-P22K1-Centro de Postre"
    / "ARCOS DORADOS PANAMA MARBELLA 101". Se cruzan por el número de loc.
    """
    import json
    if not SOW_SITES_JSON.exists():
        return {}
    sites = json.load(SOW_SITES_JSON.open())
    lookup: dict[str, dict] = {}
    for x in sites:
        text = f"{x.get('name','')} {x.get('code','')}".upper()
        m = re.search(r"\bP(\d{1,3})(K\d)?\b", text)
        if m:
            base, suffix = f"P{m.group(1)}", m.group(2)
        else:
            m = re.search(r"\b(1\d\d)\b", text)
            if not m:
                continue
            base, suffix = m.group(1), None
        entry = lookup.setdefault(base, {"lat": None, "lng": None, "name": None, "fm_ids": []})
        if x.get("order_id_fm") and str(x["order_id_fm"]) not in entry["fm_ids"]:
            entry["fm_ids"].append(str(x["order_id_fm"]))
        prefer = suffix is None or entry["name"] is None
        if prefer and (entry["lat"] is None) and x.get("lat") and x.get("lng"):
            entry["lat"], entry["lng"] = x["lat"], x["lng"]
        if suffix is None or entry["name"] is None:
            entry["name"] = re.sub(r"\s+", " ", str(x.get("name") or "")).strip() or None
    return lookup

# ---- Spanish date parser ----

SPANISH_MONTHS = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}


def parse_planned_date(val) -> datetime | None:
    """Parse PlannedDate cell. Can be:
    - datetime (Excel native date) — most rows from 2026-03 onwards
    - string in Spanish: "lunes, 12 de enero de 2026"
    - "after the previous" / None — return None (caller handles fallback)
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip().lower()
    if "after the previous" in s or s in ("", "na", "tbd"):
        return None
    # Spanish format: "lunes, 12 de enero de 2026"
    m = re.search(r"(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})", s)
    if m:
        day = int(m.group(1))
        month_name = m.group(2)
        year = int(m.group(3))
        month = SPANISH_MONTHS.get(month_name)
        if year < 2020:
            year = 2026
        if month:
            return datetime(year, month, day, tzinfo=timezone.utc)
    # Fallback: try ISO-ish
    try:
        return datetime.fromisoformat(str(val))
    except (ValueError, TypeError):
        return None


def parse_time_cell(val) -> time | None:
    """Parse a time cell (Arrival, Start, End, PlannedTime).
    Accepts datetime.time, datetime.datetime, string "HH:MM:SS", or None.
    Returns None for NA/TBD/empty/'after the previous'.
    """
    if val is None:
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    s = str(val).strip().lower()
    if s in ("", "na", "tbd", "-") or "after the previous" in s:
        return None
    # Try HH:MM:SS or HH:MM
    m = re.match(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", s)
    if m:
        h, mn, sc = int(m.group(1)), int(m.group(2)), int(m.group(3) or 0)
        try:
            return time(h, mn, sc)
        except ValueError:
            return None
    return None


def combine_date_time(d: datetime | None, t: time | None) -> datetime | None:
    """Combine a date + time into a UTC datetime.
    Note: the xlsx times are local Panamá (UTC-5). We store UTC for consistency
    with the rest of the backend. Adding 5h offset.
    """
    if d is None:
        return None
    if t is None:
        # Fallback: just the date at 00:00 UTC
        return datetime(d.year, d.month, d.day, 0, 0, 0, tzinfo=timezone.utc)
    # Panamá UTC-5 → add 5h
    naive = datetime(d.year, d.month, d.day, t.hour, t.minute, t.second)
    naive_utc = naive.replace(tzinfo=timezone.utc)
    # Adjust: the xlsx stores Panamá local, we want UTC
    from datetime import timedelta
    return naive_utc + timedelta(hours=5)


# ---- Loc.Code normalization ----

def normalize_loc_code(raw) -> tuple[str | None, str]:
    """Convert raw Loc.Code cell into (base_code, full_code).

    Examples:
    - "P22" → ("P22", "PAN-P22")
    - "P22K1" → ("P22", "PAN-P22")  (K1 means same site, additional AP)
    - "P22K2" → ("P22", "PAN-P22")
    - 100 → ("100", "PAN-100")
    - 0 → ("FINANCIAL-PARK", "PAN-FINANCIAL-PARK")  (special case)
    - "NA" / None → (None, None)  (skip)
    """
    if raw is None:
        return None, None
    s = str(raw).strip().upper()
    if s in ("", "NA", "TBD", "-"):
        return None, None
    if s == "0":
        return "FINANCIAL-PARK", "PAN-FINANCIAL-PARK"
    # Strip K-suffix (P22K1 → P22)
    base = re.sub(r"K\d+$", "", s)
    return base, f"PAN-{base}"


# ---- Asset helpers ----

def normalize_serial(raw) -> str | None:
    """Clean a serial cell. Returns None for placeholders."""
    if raw is None:
        return None
    s = str(raw).strip().strip("()").strip()
    if s.upper() in ("", "NA", "TBD", "TBC", "OMITED", "OMITTED", "-"):
        return None
    return s


def normalize_model(raw) -> str | None:
    """Clean a device model cell."""
    if raw is None:
        return None
    s = str(raw).strip()
    if s.upper() in ("", "NA", "TBD", "-"):
        return None
    return s


def asset_category_from_reference(ref: str | None, model: str | None) -> str:
    """Map xlsx reference + model into asset category.
    Cisco Meraki MX = firewall (network), MR/MS = AP/switch (network).
    All map to 'network' category in current asset model.
    """
    return "network"


# ---- Mongo upserts ----

def now_utc() -> datetime:
    return datetime.now(timezone.utc)


async def upsert_site(db, *, tenant_id: str, organization_id: str, full_code: str, doc: dict) -> str:
    key = {"code": full_code, "tenant_id": tenant_id, "organization_id": organization_id}
    existing = await db.sites.find_one(key)
    if existing:
        await db.sites.update_one({"_id": existing["_id"]}, {"$set": {**doc, "updated_at": now_utc()}})
        return str(existing["_id"])
    doc.setdefault("created_at", now_utc())
    doc.setdefault("updated_at", now_utc())
    res = await db.sites.insert_one(doc)
    return str(res.inserted_id)


async def upsert_work_order(db, *, reference: str, doc: dict) -> str:
    key = {"reference": reference, "tenant_id": doc["tenant_id"]}
    existing = await db.work_orders.find_one(key)
    if existing:
        await db.work_orders.update_one({"_id": existing["_id"]}, {"$set": {**doc, "updated_at": now_utc()}})
        return str(existing["_id"])
    doc.setdefault("created_at", now_utc())
    doc.setdefault("updated_at", now_utc())
    res = await db.work_orders.insert_one(doc)
    return str(res.inserted_id)


async def upsert_asset(db, *, tenant_id: str, serial_number: str, doc: dict) -> str:
    key = {"serial_number": serial_number, "tenant_id": tenant_id}
    existing = await db.assets.find_one(key)
    if existing:
        await db.assets.update_one({"_id": existing["_id"]}, {"$set": {**doc, "updated_at": now_utc()}})
        return str(existing["_id"])
    doc.setdefault("created_at", now_utc())
    doc.setdefault("updated_at", now_utc())
    res = await db.assets.insert_one(doc)
    return str(res.inserted_id)


async def insert_or_replace_asset_event(db, *, asset_id: str, intervention_id: str, event_type: str, doc: dict):
    """Replace any existing event for (asset, intervention, event_type) to keep idempotent."""
    await db.asset_events.delete_many({
        "asset_id": asset_id,
        "intervention_id": intervention_id,
        "event_type": event_type,
    })
    doc.setdefault("created_at", now_utc())
    doc.setdefault("updated_at", now_utc())
    await db.asset_events.insert_one(doc)


# ---- Main ----

async def main(dry_run: bool = False):
    import openpyxl

    if not XLSX_PATH.exists():
        print(f"ERROR: {XLSX_PATH} not found")
        sys.exit(1)

    print(f"Reading: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Control"]

    await connect_db()
    db = get_db()
    assert db is not None, "DB connection failed"

    # Pick first tenant
    tenant = await db.tenants.find_one({})
    if not tenant:
        print("ERROR: no tenants. Run seed_foundation first.")
        sys.exit(1)
    tenant_id = str(tenant["_id"])

    # Pick CES org + project + agreement (from seed_arcos_claro)
    ces_org = await db.organizations.find_one({"display_name": "Claro CES", "tenant_id": tenant_id})
    arcos_org = await db.organizations.find_one({"display_name": "Arcos Dorados", "tenant_id": tenant_id})
    project = await db.projects.find_one({"code": PROJECT_CODE, "tenant_id": tenant_id})
    agreement = await db.service_agreements.find_one({"contract_ref": SERVICE_AGREEMENT_REF, "tenant_id": tenant_id})
    tech_user = await db.users.find_one({"email": TECH_EMAIL})
    coord_user = await db.users.find_one({"email": COORD_EMAIL})
    sow = load_sow_lookup()

    missing = []
    if not ces_org: missing.append("CES org")
    if not arcos_org: missing.append("Arcos org")
    if not project: missing.append(f"Project {PROJECT_CODE}")
    if not agreement: missing.append(f"Agreement {SERVICE_AGREEMENT_REF}")
    if not tech_user: missing.append(f"Tech user {TECH_EMAIL}")
    if not coord_user: missing.append(f"Coord user {COORD_EMAIL}")
    if missing:
        print(f"ERROR: missing prerequisites: {missing}. Run seed_foundation + seed_arcos_claro first.")
        sys.exit(1)

    ces_id = str(ces_org["_id"])
    arcos_id = str(arcos_org["_id"])
    project_id = str(project["_id"])
    agreement_id = str(agreement["_id"])
    tech_id = str(tech_user["_id"])
    coord_id = str(coord_user["_id"])

    print(f"Tenant: {tenant_id}")
    print(f"Coordinador SRS (Andros): {coord_id} · SOW lookup: {len(sow)} locs con coordenadas/FM")
    print(f"Project: {project_id} (code: {PROJECT_CODE})")
    print(f"Tech (Agustín): {tech_id}")

    # ---- Parse all rows ----
    # Group by (loc_base, planned_date_str)
    visits = {}  # (loc_base, date_str) -> {planned_date, planned_time, arrival, real_start, end, address, rows: [...]}
    last_planned_date = None  # for "after the previous" inheritance
    last_planned_time = None

    for row_idx in range(4, ws.max_row + 1):
        cells = [ws.cell(row=row_idx, column=c).value for c in range(1, 16)]
        devices, serial, reference, planned_date_raw, planned_time_raw, arrival, real_start, end, duration, early, previous, status, address, loc_code_raw, notes = cells

        if not any(cells):
            continue

        # Skip rows with no Loc.Code or fully NA
        loc_base, full_loc_code = normalize_loc_code(loc_code_raw)
        if loc_base is None:
            print(f"  R{row_idx}: skip (no loc_code)")
            continue

        # Parse planned date — fallback to last seen if "after the previous"
        planned_date = parse_planned_date(planned_date_raw)
        if planned_date is None:
            planned_date = last_planned_date
        else:
            last_planned_date = planned_date

        if planned_date is None:
            print(f"  R{row_idx}: skip (no planned_date)")
            continue

        # Parse planned time — fallback to last seen
        planned_time = parse_time_cell(planned_time_raw)
        if planned_time is None:
            planned_time = last_planned_time
        else:
            last_planned_time = planned_time

        date_key = planned_date.strftime("%Y-%m-%d")
        visit_key = (loc_base, date_key)

        if visit_key not in visits:
            visits[visit_key] = {
                "loc_base": loc_base,
                "full_loc_code": full_loc_code,
                "planned_date": planned_date,
                "planned_time": planned_time,
                "arrival": parse_time_cell(arrival),
                "real_start": parse_time_cell(real_start),
                "end": parse_time_cell(end),
                "address": str(address).strip() if address else None,
                "status_set": set(),
                "rows": [],
            }
        v = visits[visit_key]
        v["rows"].append({
            "row_idx": row_idx,
            "model": normalize_model(devices),
            "serial": normalize_serial(serial),
            "reference": str(reference).strip() if reference and str(reference).strip().upper() not in ("NA", "TBD") else None,
            "status": str(status).strip() if status else None,
            "notes": str(notes).strip() if notes and str(notes).strip().lower() != "no comments" else None,
        })
        if status:
            v["status_set"].add(str(status).strip())
        # Update arrival/start/end if this row has them and visit doesn't
        if v["arrival"] is None: v["arrival"] = parse_time_cell(arrival)
        if v["real_start"] is None: v["real_start"] = parse_time_cell(real_start)
        if v["end"] is None: v["end"] = parse_time_cell(end)
        if v["address"] is None and address: v["address"] = str(address).strip()

    print(f"\nParsed {len(visits)} unique visits from {ws.max_row - 3} rows")

    # ---- Compute summary ----
    completed = sum(1 for v in visits.values() if "Completed" in v["status_set"])
    cancelled = sum(1 for v in visits.values() if "Cancelled" in v["status_set"] and "Completed" not in v["status_set"])
    other = len(visits) - completed - cancelled
    print(f"  Completed: {completed} · Cancelled: {cancelled} · Other: {other}")

    if dry_run:
        print("\n*** DRY RUN — no writes ***")
        print(f"{'LOC':8} {'FECHA':10} {'ESTADO':10} {'EQ':>2} {'COORD':5} {'FM':14} NOMBRE/DIRECCIÓN")
        for vk, v in sorted(visits.items(), key=lambda kv: kv[0][1]):
            info = sow.get(v["loc_base"]) or {}
            st = "closed" if "Completed" in v["status_set"] else ("cancelled" if v["status_set"] <= {"Cancelled", "NA"} else "closed")
            print(f"{v['loc_base']:8} {vk[1]:10} {st:10} {len(v['rows']):>2} {'sí' if info.get('lat') else 'no':5} {','.join(info.get('fm_ids', []))[:14]:14} {(info.get('name') or v['address'] or '')[:48]}")
        for vk, v in list(visits.items())[:5]:
            print(f"\n{vk}:")
            print(f"  scheduled: {combine_date_time(v['planned_date'], v['planned_time'])}")
            print(f"  arrival:   {combine_date_time(v['planned_date'], v['arrival'])}")
            print(f"  end:       {combine_date_time(v['planned_date'], v['end'])}")
            print(f"  address:   {v['address']}")
            print(f"  status:    {v['status_set']}")
            print(f"  rows:      {len(v['rows'])}")
            for r in v["rows"]:
                print(f"    · {r['model']} / {r['serial']} ({r['status']})")
        return

    # ---- Apply writes ----
    sla_snapshot = {
        "receive_minutes": 4320,
        "resolve_minutes": 480,
        "photos_required": "all",
        "escalation_role": "project_manager",
        "escalation_minutes": 240,
        "coverage_247": True,
        "dedicated_coordinator": True,
    }

    sites_created = 0
    wos_created = 0
    assets_created = 0
    events_created = 0

    for visit_key, v in visits.items():
        loc_base = v["loc_base"]
        full_loc_code = v["full_loc_code"]
        scheduled_at = combine_date_time(v["planned_date"], v["planned_time"])
        arrival_at = combine_date_time(v["planned_date"], v["arrival"])
        end_at = combine_date_time(v["planned_date"], v["end"])

        # Determine WO status from row statuses
        s_set = v["status_set"]
        if "Completed" in s_set:
            wo_status = "closed"
            ball_side = "srs"  # closed = SRS done
        elif s_set <= {"Cancelled", "NA"}:
            wo_status = "cancelled"
            ball_side = "srs"
        else:
            wo_status = "closed"
            ball_side = "srs"

        # 1. Upsert site (coordenadas + nombre desde la lista SOW cuando cruza)
        info = sow.get(loc_base) or {}
        fm_ids = info.get("fm_ids") or []
        site_doc = {
            "tenant_id": tenant_id,
            "organization_id": arcos_id,  # site belongs to end-client Arcos
            "code": full_loc_code,
            "name": info.get("name") or v["address"] or full_loc_code,
            "country": "PA",
            "city": "Panamá",
            "address": v["address"],
            "lat": info.get("lat"),
            "lng": info.get("lng"),
            "site_type": "retail",
            "status": "active",
            "timezone": "America/Panama",
            "notes": f"Phase II Panamá — PA-1000066. Control real Agustín 25-abr-2026. Loc: {loc_base}."
                     + (f" FM Claro: {', '.join(fm_ids)}." if fm_ids else ""),
        }
        site_id = await upsert_site(
            db, tenant_id=tenant_id, organization_id=arcos_id,
            full_code=full_loc_code, doc=site_doc,
        )
        sites_created += 1

        # 2. Upsert WO
        wo_reference = f"PAN-{loc_base}-{v['planned_date'].strftime('%Y%m%d')}"

        handshakes = []
        if arrival_at:
            handshakes.append({
                "kind": "check_in",
                "ts": arrival_at,
                "actor_user_id": tech_id,
                "notes": "Arrival registered from xlsx control",
            })
        if end_at and wo_status == "closed":
            handshakes.append({
                "kind": "resolution",
                "ts": end_at,
                "actor_user_id": tech_id,
                "notes": "End time registered from xlsx control",
            })
            handshakes.append({
                "kind": "closure",
                "ts": end_at,
                "actor_user_id": tech_id,
                "notes": "Auto-closed (Completed status in control)",
            })

        # WO description
        device_summary = ", ".join(
            f"{r['model'] or '?'} ({r['reference'] or '?'})"
            for r in v["rows"] if r["model"]
        )
        notes_summary = " | ".join(r["notes"] for r in v["rows"] if r["notes"])

        base_ts = scheduled_at or now_utc()
        status_timestamps = {st: base_ts for st in PANAMA_STATUS_ORDER}
        onsite_at = combine_date_time(v["planned_date"], v["real_start"]) or arrival_at
        if arrival_at:
            status_timestamps["en_route"] = arrival_at
        if onsite_at:
            status_timestamps["on_site"] = onsite_at
        if wo_status == "closed":
            status_timestamps["resolved"] = end_at or onsite_at or base_ts
            status_timestamps["closed"] = end_at or onsite_at or base_ts
        else:
            status_timestamps["cancelled"] = base_ts

        wo_doc = {
            "tenant_id": tenant_id,
            "organization_id": ces_id,  # client facturador
            "site_id": site_id,
            "service_agreement_id": agreement_id,
            "project_id": project_id,
            "reference": wo_reference,
            "title": f"PAN-{loc_base} {info.get('name') or v['address'] or ''} — Instalación SDWAN".strip(),
            "description": (f"Equipos: {device_summary}. {notes_summary}".strip()
                            + (f" FM Claro: {', '.join(fm_ids)}." if fm_ids else "")),
            "severity": "normal",
            "status": wo_status,
            "shield_level": "gold",
            "sla_snapshot": sla_snapshot,
            "ball_in_court": {
                "side": ball_side,
                "actor_user_id": coord_id,
                "since": end_at or scheduled_at or now_utc(),
                "reason": "Importado del control real de Agustín 25-abr-2026",
            },
            "srs_coordinator_user_id": coord_id,
            "status_timestamps": status_timestamps,
            "scheduled_at": scheduled_at,
            "deadline_resolve_at": (scheduled_at + timedelta(minutes=sla_snapshot["resolve_minutes"])) if scheduled_at else None,
            "assigned_tech_user_id": tech_id,
            "handshakes": handshakes,
            "pre_flight_checklist": {},
            "after_hours": False,
            "closed_at": end_at if wo_status == "closed" else None,
            "cancelled_at": scheduled_at if wo_status == "cancelled" else None,
            "cancel_reason": notes_summary if wo_status == "cancelled" else None,
        }
        wo_id = await upsert_work_order(db, reference=wo_reference, doc=wo_doc)
        wos_created += 1

        # 3. Per row: upsert asset + asset_event (if has serial and is completed)
        for r in v["rows"]:
            if not r["serial"] or not r["model"]:
                continue
            if wo_status != "closed":
                continue  # don't register install events for cancelled visits

            asset_doc = {
                "tenant_id": tenant_id,
                "organization_id": arcos_id,  # asset belongs to end-client
                "serial_number": r["serial"],
                "category": asset_category_from_reference(r["reference"], r["model"]),
                "make": "Cisco Meraki",
                "model": r["model"],
                "current_site_id": site_id,
                "status": "active",
                "lifecycle_stage": "deployed",
                "ownership": {"type": "client_owned"},
                "notes": f"Installed {scheduled_at.date() if scheduled_at else '?'} per Agustín control. Reference: {r['reference'] or '?'}.",
            }
            asset_id = await upsert_asset(db, tenant_id=tenant_id, serial_number=r["serial"], doc=asset_doc)
            assets_created += 1

            event_doc = {
                "tenant_id": tenant_id,
                "asset_id": asset_id,
                "event_type": "installed",
                "intervention_id": wo_id,
                "performed_by": tech_id,
                "site_id": site_id,
                "ts": end_at or scheduled_at or now_utc(),
                "data": {
                    "reference": r["reference"],
                    "model": r["model"],
                    "serial": r["serial"],
                },
                "notes": r["notes"],
                "visibility": "public",
            }
            await insert_or_replace_asset_event(
                db, asset_id=asset_id, intervention_id=wo_id,
                event_type="installed", doc=event_doc,
            )
            events_created += 1

    print(f"\n=== DONE ===")
    print(f"Sites upserted: {sites_created}")
    print(f"Work orders upserted: {wos_created}")
    print(f"Assets upserted: {assets_created}")
    print(f"Asset events: {events_created}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Parse + summary, no DB writes")
    args = parser.parse_args()
    asyncio.run(main(dry_run=args.dry_run))
